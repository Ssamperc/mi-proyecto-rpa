import pandas as pd
import re
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, Reference
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import os

def validar_email(email):
    patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(patron, str(email)))

def procesar_pedidos(ruta_archivo):
    tiempo_inicio = datetime.now()
    
    estadisticas = {
        'total_pedidos': 0,
        'pedidos_validos': 0,
        'pedidos_invalidos': 0,
        'registrados_sag': 0,
        'fallidos_sag': 0,
        'correos_enviados': 0,
        'correos_fallidos': 0,
        'tasa_validacion': 0,
        'tasa_registro_sag': 0,
        'tiempo_procesamiento': '',
        'errores': [],
        'procesamiento_exitoso': False
    }
    
    output_dir = 'output'
    os.makedirs(output_dir, exist_ok=True)
    
    for filename in os.listdir(output_dir):
        file_path = os.path.join(output_dir, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
    
    try:
        if ruta_archivo.endswith('.csv'):
            df = pd.read_csv(ruta_archivo)
        else:
            df = pd.read_excel(ruta_archivo)
        
        estadisticas['total_pedidos'] = len(df)
        
        columnas_requeridas = ['ID_Pedido', 'Fecha_Pedido', 'Nombre_Cliente', 'Correo_Cliente', 
                               'Direccion_Envio', 'SKU', 'Nombre_Producto', 'Cantidad', 
                               'Precio_Unitario', 'Valor_Total']
        
        for col in columnas_requeridas:
            if col not in df.columns:
                estadisticas['errores'].append(f"Columna requerida '{col}' no encontrada en el archivo")
                return estadisticas
        
        df['Estado_Validacion'] = 'Válido'
        df['Motivo_Error'] = ''
        
        errores_detallados = []
        
        for idx, row in df.iterrows():
            errores_fila = []
            
            if pd.isna(row['ID_Pedido']) or str(row['ID_Pedido']).strip() == '':
                errores_fila.append('ID_Pedido vacío')
            
            if pd.isna(row['Fecha_Pedido']) or str(row['Fecha_Pedido']).strip() == '':
                errores_fila.append('Fecha_Pedido vacía')
            
            if pd.isna(row['Nombre_Cliente']) or str(row['Nombre_Cliente']).strip() == '':
                errores_fila.append('Nombre_Cliente vacío')
            
            if pd.isna(row['Correo_Cliente']) or not validar_email(row['Correo_Cliente']):
                errores_fila.append('Correo_Cliente inválido')
            
            if pd.isna(row['Direccion_Envio']) or str(row['Direccion_Envio']).strip() == '':
                errores_fila.append('Direccion_Envio vacía')
            
            if pd.isna(row['SKU']) or str(row['SKU']).strip() == '':
                errores_fila.append('SKU vacío')
            
            if pd.isna(row['Nombre_Producto']) or str(row['Nombre_Producto']).strip() == '':
                errores_fila.append('Nombre_Producto vacío')
            
            try:
                cantidad = float(row['Cantidad'])
                if cantidad <= 0:
                    errores_fila.append('Cantidad debe ser mayor a 0')
            except (ValueError, TypeError):
                errores_fila.append('Cantidad no es numérica')
            
            try:
                precio_unitario = float(row['Precio_Unitario'])
                if precio_unitario <= 0:
                    errores_fila.append('Precio_Unitario debe ser mayor a 0')
            except (ValueError, TypeError):
                errores_fila.append('Precio_Unitario no es numérico')
            
            try:
                valor_total = float(row['Valor_Total'])
                if valor_total <= 0:
                    errores_fila.append('Valor_Total debe ser mayor a 0')
            except (ValueError, TypeError):
                errores_fila.append('Valor_Total no es numérico')
            
            if errores_fila:
                df.at[idx, 'Estado_Validacion'] = 'Inválido'
                df.at[idx, 'Motivo_Error'] = '; '.join(errores_fila)
                estadisticas['pedidos_invalidos'] += 1
                
                errores_detallados.append({
                    'ID_Pedido': row['ID_Pedido'],
                    'Nombre_Cliente': row['Nombre_Cliente'],
                    'Errores': '; '.join(errores_fila)
                })
            else:
                estadisticas['pedidos_validos'] += 1
        
        pedidos_validos = df[df['Estado_Validacion'] == 'Válido'].copy()
        pedidos_invalidos = df[df['Estado_Validacion'] == 'Inválido'].copy()
        
        if estadisticas['pedidos_validos'] == 0:
            estadisticas['errores'].append('No hay pedidos válidos para procesar. Todos los pedidos tienen errores de validación.')
            if len(pedidos_invalidos) > 0:
                pedidos_invalidos.to_excel(f'{output_dir}/ErroresRPA.xlsx', index=False)
            return estadisticas
        
        pedidos_validos.to_excel(f'{output_dir}/PedidosValidados.xlsx', index=False)
        
        pedidos_invalidos.to_excel(f'{output_dir}/ErroresRPA.xlsx', index=False)
        
        pedidos_validos['Estado_Registro'] = 'Registrado'
        pedidos_validos['Fecha_Registro'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        if len(pedidos_validos) > 1:
            indice_fallido = random.randint(0, len(pedidos_validos) - 1)
            pedidos_validos.at[pedidos_validos.index[indice_fallido], 'Estado_Registro'] = 'Fallido'
            estadisticas['fallidos_sag'] = 1
            estadisticas['registrados_sag'] = len(pedidos_validos) - 1
        else:
            estadisticas['registrados_sag'] = len(pedidos_validos)
            estadisticas['fallidos_sag'] = 0
        
        pedidos_validos.to_excel(f'{output_dir}/PedidosRegistradosSAG.xlsx', index=False)
        
        log_correos = []
        for idx, row in pedidos_validos.iterrows():
            if row['Estado_Registro'] == 'Registrado':
                log_correos.append(
                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
                    f"Destinatario: {row['Correo_Cliente']} | "
                    f"Asunto: Confirmación de Pedido #{row['ID_Pedido']} | "
                    f"Estado: ENVIADO\n"
                )
                estadisticas['correos_enviados'] += 1
            else:
                log_correos.append(
                    f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] "
                    f"Destinatario: {row['Correo_Cliente']} | "
                    f"Asunto: Confirmación de Pedido #{row['ID_Pedido']} | "
                    f"Estado: FALLIDO - Error en registro SAG\n"
                )
                estadisticas['correos_fallidos'] += 1
        
        with open(f'{output_dir}/LogCorreos.txt', 'w', encoding='utf-8') as f:
            f.writelines(log_correos)
        
        wb_correos = Workbook()
        ws_correos = wb_correos.active
        ws_correos.title = "Resumen de Correos"
        
        ws_correos['A1'] = 'RESUMEN DE NOTIFICACIONES POR CORREO'
        ws_correos['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws_correos['A1'].fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
        ws_correos.merge_cells('A1:D1')
        ws_correos['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_correos['A3'] = 'ID Pedido'
        ws_correos['B3'] = 'Destinatario'
        ws_correos['C3'] = 'Estado Envío'
        ws_correos['D3'] = 'Fecha Envío'
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws_correos[cell].font = Font(bold=True, color='FFFFFF')
            ws_correos[cell].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        fila = 4
        for idx, row in pedidos_validos.iterrows():
            estado_envio = 'ENVIADO' if row['Estado_Registro'] == 'Registrado' else 'FALLIDO'
            ws_correos[f'A{fila}'] = row['ID_Pedido']
            ws_correos[f'B{fila}'] = row['Correo_Cliente']
            ws_correos[f'C{fila}'] = estado_envio
            ws_correos[f'D{fila}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            if estado_envio == 'ENVIADO':
                ws_correos[f'C{fila}'].fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')
                ws_correos[f'C{fila}'].font = Font(color='FFFFFF', bold=True)
            else:
                ws_correos[f'C{fila}'].fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
                ws_correos[f'C{fila}'].font = Font(color='FFFFFF', bold=True)
            
            fila += 1
        
        ws_correos.column_dimensions['A'].width = 15
        ws_correos.column_dimensions['B'].width = 30
        ws_correos.column_dimensions['C'].width = 15
        ws_correos.column_dimensions['D'].width = 20
        
        wb_correos.save(f'{output_dir}/ResumenCorreos.xlsx')
        
        wb_reporte = Workbook()
        ws_reporte = wb_reporte.active
        ws_reporte.title = "Reporte de Procesamiento"
        
        ws_reporte['A1'] = 'REPORTE DE PROCESAMIENTO RPA'
        ws_reporte['A1'].font = Font(size=14, bold=True)
        ws_reporte.merge_cells('A1:B1')
        
        ws_reporte['A3'] = 'Fecha de Procesamiento:'
        ws_reporte['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws_reporte['A5'] = 'Métrica'
        ws_reporte['B5'] = 'Valor'
        ws_reporte['A5'].font = Font(bold=True)
        ws_reporte['B5'].font = Font(bold=True)
        
        metricas = [
            ('Total de Pedidos Recibidos', estadisticas['total_pedidos']),
            ('Pedidos Válidos', estadisticas['pedidos_validos']),
            ('Pedidos Inválidos', estadisticas['pedidos_invalidos']),
            ('Registrados en SAG', estadisticas['registrados_sag']),
            ('Fallidos en SAG', estadisticas['fallidos_sag']),
            ('Correos Enviados', estadisticas['correos_enviados']),
            ('Correos Fallidos', estadisticas['correos_fallidos'])
        ]
        
        fila = 6
        for metrica, valor in metricas:
            ws_reporte[f'A{fila}'] = metrica
            ws_reporte[f'B{fila}'] = valor
            fila += 1
        
        wb_reporte.save(f'{output_dir}/ReporteProcesados.xlsx')
        
        wb_dashboard = Workbook()
        ws_dashboard = wb_dashboard.active
        ws_dashboard.title = "Dashboard RPA"
        
        ws_dashboard['A1'] = 'DASHBOARD RPA - SISTEMA DE PROCESAMIENTO DE PEDIDOS'
        ws_dashboard['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws_dashboard['A1'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
        ws_dashboard.merge_cells('A1:D1')
        ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        ws_dashboard['A3'] = 'KPI'
        ws_dashboard['B3'] = 'Valor'
        ws_dashboard['C3'] = 'Porcentaje'
        ws_dashboard['D3'] = 'Estado'
        
        for cell in ['A3', 'B3', 'C3', 'D3']:
            ws_dashboard[cell].font = Font(bold=True, color='FFFFFF')
            ws_dashboard[cell].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        tasa_validacion = (estadisticas['pedidos_validos'] / estadisticas['total_pedidos'] * 100) if estadisticas['total_pedidos'] > 0 else 0
        tasa_registro_sag = (estadisticas['registrados_sag'] / estadisticas['pedidos_validos'] * 100) if estadisticas['pedidos_validos'] > 0 else 0
        
        estadisticas['tasa_validacion'] = round(tasa_validacion, 2)
        estadisticas['tasa_registro_sag'] = round(tasa_registro_sag, 2)
        
        kpis = [
            ('Total Pedidos Recibidos', estadisticas['total_pedidos'], '100%', '✓'),
            ('Pedidos Válidos', estadisticas['pedidos_validos'], f"{tasa_validacion:.1f}%", '✓' if tasa_validacion > 80 else '⚠'),
            ('Pedidos Inválidos', estadisticas['pedidos_invalidos'], f"{100-tasa_validacion:.1f}%", '⚠' if estadisticas['pedidos_invalidos'] > 0 else '✓'),
            ('Registrados en SAG', estadisticas['registrados_sag'], f"{tasa_registro_sag:.1f}%", '✓'),
            ('Fallidos en SAG', estadisticas['fallidos_sag'], f"{100-tasa_registro_sag:.1f}%", '⚠' if estadisticas['fallidos_sag'] > 0 else '✓'),
            ('Correos Enviados', estadisticas['correos_enviados'], f"{(estadisticas['correos_enviados']/estadisticas['total_pedidos']*100) if estadisticas['total_pedidos'] > 0 else 0:.1f}%", '✓'),
        ]
        
        fila = 4
        for kpi, valor, porcentaje, estado in kpis:
            ws_dashboard[f'A{fila}'] = kpi
            ws_dashboard[f'B{fila}'] = valor
            ws_dashboard[f'C{fila}'] = porcentaje
            ws_dashboard[f'D{fila}'] = estado
            fila += 1
        
        ws_dashboard.column_dimensions['A'].width = 25
        ws_dashboard.column_dimensions['B'].width = 12
        ws_dashboard.column_dimensions['C'].width = 15
        ws_dashboard.column_dimensions['D'].width = 10
        
        wb_dashboard.save(f'{output_dir}/DashboardRPA.xlsx')
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
        
        labels_validacion = ['Válidos', 'Inválidos']
        sizes_validacion = [estadisticas['pedidos_validos'], estadisticas['pedidos_invalidos']]
        colors_validacion = ['#28a745', '#dc3545']
        
        ax1.pie(sizes_validacion, labels=labels_validacion, autopct='%1.1f%%', 
                colors=colors_validacion, startangle=90)
        ax1.set_title('Estado de Validación de Pedidos', fontsize=14, fontweight='bold')
        
        categorias = ['Total\nPedidos', 'Válidos', 'Inválidos', 'Registrados\nSAG', 'Correos\nEnviados']
        valores = [
            estadisticas['total_pedidos'],
            estadisticas['pedidos_validos'],
            estadisticas['pedidos_invalidos'],
            estadisticas['registrados_sag'],
            estadisticas['correos_enviados']
        ]
        colors_barras = ['#007bff', '#28a745', '#dc3545', '#17a2b8', '#ffc107']
        
        bars = ax2.bar(categorias, valores, color=colors_barras)
        ax2.set_title('Resumen de Procesamiento', fontsize=14, fontweight='bold')
        ax2.set_ylabel('Cantidad', fontsize=12)
        
        for bar in bars:
            height = bar.get_height()
            ax2.text(bar.get_x() + bar.get_width()/2., height,
                    f'{int(height)}',
                    ha='center', va='bottom', fontsize=10, fontweight='bold')
        
        plt.tight_layout()
        plt.savefig(f'{output_dir}/graficos_dashboard.png', dpi=150, bbox_inches='tight')
        plt.close()
        
        tiempo_fin = datetime.now()
        tiempo_transcurrido = tiempo_fin - tiempo_inicio
        estadisticas['tiempo_procesamiento'] = str(tiempo_transcurrido.total_seconds()) + ' segundos'
        
        estadisticas['errores'] = errores_detallados
        
        archivos_requeridos = [
            'PedidosValidados.xlsx',
            'ErroresRPA.xlsx',
            'PedidosRegistradosSAG.xlsx',
            'ReporteProcesados.xlsx',
            'DashboardRPA.xlsx',
            'ResumenCorreos.xlsx',
            'LogCorreos.txt',
            'graficos_dashboard.png'
        ]
        
        archivos_faltantes = []
        for archivo in archivos_requeridos:
            ruta_completa = os.path.join(output_dir, archivo)
            if not os.path.exists(ruta_completa):
                archivos_faltantes.append(archivo)
        
        if archivos_faltantes:
            estadisticas['errores'].append(f'Archivos no generados correctamente: {", ".join(archivos_faltantes)}')
            for filename in os.listdir(output_dir):
                file_path = os.path.join(output_dir, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            return estadisticas
        
        estadisticas['procesamiento_exitoso'] = True
        
        return estadisticas
        
    except Exception as e:
        estadisticas['errores'].append(f'Error general: {str(e)}')
        for filename in os.listdir(output_dir):
            file_path = os.path.join(output_dir, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
        return estadisticas
